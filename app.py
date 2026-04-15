# =====================================================================
# ADIDAS FINANCIAL ANALYSIS DASHBOARD
# =====================================================================
# Streamlit + Plotly single-file dashboard.
# Data source: Balance Sheet 2021-2024 (values in €K, displayed in €M).
# =====================================================================

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import numpy as np
from pathlib import Path

# ─── Page configuration ─────────────────────────────────────────────
st.set_page_config(
    layout="wide",
    page_title="Adidas Financial Dashboard",
    page_icon="👟",
    initial_sidebar_state="expanded",
)

# ─── Constants ───────────────────────────────────────────────────────
FILE_PATH = Path(__file__).parent / "Easy-to-use simplified balance sheet Financial diagnosis_Adidas.xlsx"
YEARS = [2021, 2022, 2023, 2024]
YEARS_STR = [str(y) for y in YEARS]

# Colour palette
BG      = "#0d1117"
CARD_BG = "#161b22"
BORDER  = "#30363d"
TXT     = "#ffffff"
TXT2    = "#8b949e"
BLUE    = "#004B87"
GREEN   = "#00A651"
ORANGE  = "#FF6B00"
RED     = "#E31937"
PALETTE = [BLUE, GREEN, ORANGE, RED, "#9B59B6", "#1ABC9C", "#F39C12", "#2ECC71"]

# Ordered line-item lists (match Excel row order after renaming)
ASSET_ITEMS = [
    "Intangible assets", "Property, plant and equipment",
    "Other non-current assets", "Total non-current assets",
    "Inventories", "Accounts receivable", "Other current assets",
    "Cash and cash equivalents", "Total current assets", "Total assets",
]
EQLIAB_ITEMS = [
    "Issued share capital", "Capital & Other reserves",
    "Equity portion of convertible debt", "Retained earnings", "Total equity",
    "Borrowings non-current", "Other non-current liabilities",
    "Total non-current liabilities", "Accounts payable", "Bank overdraft",
    "Borrowings current", "Other current liabilities",
    "Total current liabilities", "Total equity and liabilities",
]
ALL_ITEMS = ASSET_ITEMS + EQLIAB_ITEMS

# Short names for tight chart labels
_SN = {
    "Property, plant and equipment": "PP&E",
    "Other non-current assets": "Other NCA",
    "Total non-current assets": "Total NCA",
    "Cash and cash equivalents": "Cash & Equiv.",
    "Total current assets": "Total CA",
    "Accounts receivable": "Accounts Recv.",
    "Other current assets": "Other CA",
    "Issued share capital": "Share Capital",
    "Capital & Other reserves": "Capital & Reserves",
    "Equity portion of convertible debt": "Conv. Debt Equity",
    "Borrowings non-current": "Borrowings (NC)",
    "Other non-current liabilities": "Other NCL",
    "Total non-current liabilities": "Total NCL",
    "Accounts payable": "Accounts Pay.",
    "Borrowings current": "Borrowings (Curr.)",
    "Other current liabilities": "Other CL",
    "Total current liabilities": "Total CL",
    "Total equity and liabilities": "Total E&L",
}

def sn(name):
    return _SN.get(name, name)

# ─── Scenario definitions ───────────────────────────────────────────
SCENARIOS = {
    1: ("Ideal", "✅", GREEN,
        "WC > 0 · WCN > 0 · NC > 0",
        "Healthy cycle. Long-term resources cover fixed assets "
        "with extra cash for operations."),
    2: ("Risky but Common", "⚠️", ORANGE,
        "WC > 0 · WCN > 0 · NC < 0",
        "Stable long-term, but short-term liquidity tension. "
        "Operating cycle too heavy, needs short-term bank loans."),
    3: ("Excellent", "🌟", "#00D4AA",
        "WC > 0 · WCN < 0 · NC > 0",
        "Perfect situation (common in retail). Negative WCN means "
        "operations generate cash. Solid long-term, excess liquidity."),
    4: ("Paradoxical", "🔄", "#FFD700",
        "WC < 0 · WCN < 0 · NC > 0",
        "Risky structure (fixed assets financed by short-term debt), "
        "but business generates enough daily cash to survive. "
        "If activity slows, massive liquidity problem."),
    5: ("Risky", "🔴", RED,
        "WC < 0 · WCN < 0 · NC < 0",
        "Bad structure + cash consumption. Extreme financial tension, "
        "over-reliant on short-term debt."),
    6: ("Dangerous", "💀", "#FF0000",
        "WC < 0 · WCN > 0 · NC < 0",
        "Worst situation. Weak structure, cash-consuming operating "
        "cycle, severe liquidity distress. High bankruptcy risk."),
}

# ─── CSS injection ───────────────────────────────────────────────────
def inject_css():
    st.markdown(f"""<style>
    .stApp {{ background:{BG}; color:{TXT}; }}
    [data-testid="stSidebar"] {{ background:#0a0e14; border-right:1px solid {BORDER}; }}
    [data-testid="stSidebar"] * {{ color:{TXT}!important; }}
    .card {{ background:{CARD_BG}; border:1px solid {BORDER}; border-radius:12px;
             padding:18px 20px; margin-bottom:10px; }}
    .card h4 {{ color:{TXT2}; font-size:.82rem; margin:0 0 4px; font-weight:500; }}
    .card .val {{ font-size:1.85rem; font-weight:700; line-height:1.2; }}
    .card .sub {{ color:{TXT2}; font-size:.78rem; margin-top:4px; }}
    .scenario {{ background:{CARD_BG}; border:1px solid {BORDER}; border-radius:12px;
                 padding:14px 18px; margin-bottom:10px; border-left:4px solid; }}
    .ipanel {{ background:{CARD_BG}; border:1px solid {BORDER}; border-radius:10px;
               padding:16px; margin:8px 0; }}
    .ipanel .formula {{ font-family:'Courier New',monospace; background:{BG}; padding:8px 12px;
                        border-radius:6px; margin:6px 0; color:{BLUE}; font-size:.85rem; }}
    .ipanel .verdict {{ margin-top:8px; padding:10px; border-radius:6px; font-weight:600; }}
    .shdr {{ font-size:1.15rem; font-weight:700; margin:18px 0 8px;
             padding-bottom:6px; border-bottom:2px solid {BLUE}; color:{TXT}; }}
    footer {{ visibility:hidden; }}
    </style>""", unsafe_allow_html=True)


# ─── Data loading ────────────────────────────────────────────────────
@st.cache_data
def load_data():
    """Load balance sheet from Excel. Returns DataFrame with values in €K."""
    import openpyxl
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb["Balance Sheet"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    header = rows[0]
    ycols = {v: i for i, v in enumerate(header) if v in YEARS}

    data = {}
    bcount = 0
    for r in rows[1:]:
        lbl = r[1]
        if lbl is None:
            continue
        if "Borrowings" in str(lbl) and "similar" in str(lbl):
            bcount += 1
            lbl = "Borrowings non-current" if bcount == 1 else "Borrowings current"
        vals = {y: (r[c] if r[c] is not None else 0) for y, c in ycols.items()}
        data[lbl] = vals

    df = pd.DataFrame(data).T[sorted(ycols.keys())].astype(float)
    return df  # values in €K


# ─── Computations ────────────────────────────────────────────────────
def _g(df, label, yr):
    try:
        return float(df.loc[label, yr])
    except (KeyError, TypeError):
        return 0.0


def compute(df):
    """Return {year: {metric_name: value}} for every computed metric."""
    out = {}
    for yr in YEARS:
        eq   = _g(df, "Total equity", yr)
        ncl  = _g(df, "Total non-current liabilities", yr)
        nca  = _g(df, "Total non-current assets", yr)
        ca   = _g(df, "Total current assets", yr)
        cl   = _g(df, "Total current liabilities", yr)
        ta   = _g(df, "Total assets", yr)
        inv  = _g(df, "Inventories", yr)
        ar   = _g(df, "Accounts receivable", yr)
        oca  = _g(df, "Other current assets", yr)
        cash = _g(df, "Cash and cash equivalents", yr)
        ap   = _g(df, "Accounts payable", yr)
        ocl  = _g(df, "Other current liabilities", yr)
        bnc  = _g(df, "Borrowings non-current", yr)
        bc   = _g(df, "Borrowings current", yr)
        bo   = _g(df, "Bank overdraft", yr)

        wc  = eq + ncl - nca
        wcn = (inv + ar + oca) - (ap + ocl)
        nc  = wc - wcn
        nd  = bnc + bc - cash
        ce  = nca + wcn
        ic  = eq + nd
        debt = bnc + bc

        cr    = ca / cl if cl else 0
        qr    = (ca - inv) / cl if cl else 0
        cashr = cash / cl if cl else 0
        em    = ta / eq if eq else 0
        tdr   = debt / ta if ta else 0
        de    = debt / eq if eq else 0

        # Scenario classification
        s = (1 if wc > 0 and wcn > 0 and nc > 0 else
             2 if wc > 0 and wcn > 0 and nc < 0 else
             3 if wc > 0 and wcn < 0 and nc > 0 else
             4 if wc < 0 and wcn < 0 and nc > 0 else
             5 if wc < 0 and wcn < 0 and nc < 0 else
             6 if wc < 0 and wcn > 0 and nc < 0 else 0)

        out[yr] = {
            "WC": wc, "WCN": wcn, "NC": nc,
            "NC_verify": cash - bc - bo,
            "Net Debt": nd, "Capital Employed": ce, "Invested Capital": ic,
            "Current Ratio": cr, "Quick Ratio": qr, "Cash Ratio": cashr,
            "Equity Multiplier": em, "Total Debt Ratio": tdr, "Debt-to-Equity": de,
            "Scenario": s,
        }
    return out


# ─── Rating functions ────────────────────────────────────────────────
def rate_cr(v):
    if v < 1:    return "🔴", RED, "Liquidity concern — current liabilities exceed current assets"
    if v < 1.5:  return "🟡", ORANGE, "Below ideal range — adequate but tight"
    if v <= 2:   return "🟢", GREEN, "Healthy — adequate liquidity for short-term obligations"
    if v <= 3:   return "🟢", GREEN, "Good liquidity position"
    return "🟡", ORANGE, "Inefficient — too much working capital tied up"

def rate_qr(v):
    if v < 1:    return "🔴", RED, "Liquidity risk — heavily dependent on selling inventory"
    if v <= 1.5: return "🟢", GREEN, "Strong, safe liquidity without relying on inventory"
    if v <= 2:   return "🟢", GREEN, "Very good quick liquidity"
    return "🟡", ORANGE, "Very conservative — could mean under-investing in growth"

def rate_cashr(v):
    if v < 0.5:  return "🟡", ORANGE, "Reliance on inventory/receivables to pay short-term debts"
    if v <= 1:   return "🟢", GREEN, "Ideal cash position relative to current liabilities"
    return "🟡", ORANGE, "Extremely conservative — excess cash may be idle"

def rate_tdr(v):
    if v <= 0.2: return "🟢", GREEN, "Very low debt"
    if v <= 0.4: return "🟢", GREEN, "Low debt — healthy capital structure"
    if v <= 0.6: return "🟡", ORANGE, "Acceptable for most industries"
    if v <= 0.8: return "🔴", RED, "High debt load — caution advised"
    return "🔴🔴", RED, "Very high leverage — significant solvency risk"

def rate_de(v):
    if v <= 0.5: return "🟢", GREEN, "Very healthy — low leverage"
    if v <= 1:   return "🟢", GREEN, "Healthy, balanced capital structure"
    if v <= 1.5: return "🟡", ORANGE, "Acceptable leverage"
    return "🔴", RED, "High leverage — potential solvency concerns"

def rate_sign(v, label=""):
    if v > 0: return "🟢", GREEN, f"Positive {label}".strip()
    if v < 0: return "🔴", RED, f"Negative {label}".strip()
    return "⚪", TXT2, f"Zero {label}".strip()


# ─── Formatting helpers ──────────────────────────────────────────────
def fm(v_k):
    """€K value → €M display string."""
    return f"€{v_k / 1000:,.1f}M"

def fr(v):
    """Ratio → 2 decimal places."""
    return f"{v:.2f}"

def fp(v):
    """Percentage with 1 decimal."""
    return f"{v:.1f}%"

def yoy_pct(curr, prev):
    if prev == 0:
        return None
    return ((curr - prev) / abs(prev)) * 100

def arrow_html(val):
    if val is None:
        return f'<span style="color:{TXT2};">N/A</span>'
    color = GREEN if val >= 0 else RED
    sym = "↑" if val >= 0 else "↓"
    return f'<span style="color:{color};font-weight:600;">{sym}&nbsp;{abs(val):.1f}%</span>'


# ─── Chart helper ────────────────────────────────────────────────────
def dark_fig(fig, height=420):
    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor=BG, plot_bgcolor=BG,
        font=dict(family="sans-serif", color=TXT, size=12),
        legend=dict(bgcolor="rgba(0,0,0,0)", bordercolor=BORDER),
        margin=dict(l=50, r=30, t=50, b=40),
        height=height,
    )
    fig.update_xaxes(gridcolor=BORDER, linecolor=BORDER)
    fig.update_yaxes(gridcolor=BORDER, linecolor=BORDER)
    return fig


# ─── UI components ───────────────────────────────────────────────────
def card_html(title, value, color=TXT, subtitle=""):
    st.markdown(
        f'<div class="card"><h4>{title}</h4>'
        f'<div class="val" style="color:{color};">{value}</div>'
        f'<div class="sub">{subtitle}</div></div>',
        unsafe_allow_html=True,
    )

def scenario_card_html(year, case_num):
    label, icon, color, signs, interp = SCENARIOS.get(case_num, ("Unknown", "❓", TXT2, "", ""))
    st.markdown(
        f'<div class="scenario" style="border-left-color:{color};">'
        f'<div style="font-size:.78rem;color:{TXT2};">FY {year}</div>'
        f'<div style="font-size:1.3rem;font-weight:700;color:{color};">'
        f'{icon} Case {case_num}: {label}</div>'
        f'<div style="font-size:.72rem;color:{TXT2};margin:4px 0;">{signs}</div>'
        f'<div style="font-size:.82rem;color:{TXT};margin-top:6px;">{interp}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

def interp_panel_html(title, formula_text, value_str, thresholds_html, verdict, vcolor):
    st.markdown(
        f'<div class="ipanel">'
        f'<div style="font-weight:700;font-size:1rem;margin-bottom:6px;">{title}</div>'
        f'<div class="formula">{formula_text}</div>'
        f'<div style="margin:6px 0;"><strong>Value:</strong> '
        f'<span style="color:{vcolor};font-size:1.2rem;font-weight:700;">{value_str}</span></div>'
        f'<div style="margin:6px 0;font-size:.85rem;">{thresholds_html}</div>'
        f'<div class="verdict" style="background:{vcolor}18;border-left:3px solid {vcolor};color:{vcolor};">'
        f'{verdict}</div></div>',
        unsafe_allow_html=True,
    )

def section(title):
    st.markdown(f'<div class="shdr">{title}</div>', unsafe_allow_html=True)


# =====================================================================
# PAGE 1 — Executive Summary
# =====================================================================
def page_summary(df, M):
    st.title("👟 Adidas — Executive Summary")

    # ── Row 1: Scenario snapshot per year ──
    section("Health Snapshot by Year")
    cols = st.columns(4)
    for i, yr in enumerate(YEARS):
        with cols[i]:
            scenario_card_html(yr, M[yr]["Scenario"])

    # ── Row 2: Key metrics 2024 ──
    section("Key Metrics — FY 2024")
    m = M[2024]
    cols = st.columns(4)
    with cols[0]:
        _, c, sub = rate_sign(m["WC"], "Working Capital")
        card_html("Working Capital", fm(m["WC"]), c, sub)
    with cols[1]:
        _, c, sub = rate_sign(m["NC"], "Net Cash")
        card_html("Net Cash", fm(m["NC"]), c, sub)
    with cols[2]:
        emoji, c, sub = rate_cr(m["Current Ratio"])
        card_html("Current Ratio", f"{emoji} {fr(m['Current Ratio'])}", c, sub)
    with cols[3]:
        emoji, c, sub = rate_de(m["Debt-to-Equity"])
        card_html("Debt-to-Equity", f"{emoji} {fr(m['Debt-to-Equity'])}", c, sub)

    # ── Row 3: WC / WCN / NC evolution chart ──
    section("WC / WCN / NC Evolution (€M)")
    fig = go.Figure()
    for name, key, color in [("Working Capital", "WC", BLUE),
                              ("WCN", "WCN", ORANGE),
                              ("Net Cash", "NC", GREEN)]:
        vals = [M[y][key] / 1000 for y in YEARS]
        fig.add_trace(go.Scatter(
            x=YEARS_STR, y=vals, name=name, mode="lines+markers",
            line=dict(color=color, width=3), marker=dict(size=9),
        ))
    fig.add_hline(y=0, line_dash="dash", line_color=TXT2, opacity=0.4)
    fig.update_layout(yaxis_title="€M", xaxis_title="Year")
    st.plotly_chart(dark_fig(fig), use_container_width=True)

    # ── Row 4: Summary table of ALL computed metrics ──
    section("All Computed Metrics")
    agg_keys = ["WC", "WCN", "NC", "Net Debt", "Capital Employed", "Invested Capital"]
    liq_keys = ["Current Ratio", "Quick Ratio", "Cash Ratio"]
    sol_keys = ["Equity Multiplier", "Total Debt Ratio", "Debt-to-Equity"]

    def _color_for(key, val):
        if key == "Current Ratio":    _, c, _ = rate_cr(val); return c
        if key == "Quick Ratio":      _, c, _ = rate_qr(val); return c
        if key == "Cash Ratio":       _, c, _ = rate_cashr(val); return c
        if key == "Total Debt Ratio": _, c, _ = rate_tdr(val); return c
        if key == "Debt-to-Equity":   _, c, _ = rate_de(val); return c
        if key == "Equity Multiplier": return BLUE
        _, c, _ = rate_sign(val); return c

    def _rows(keys, fmt):
        h = ""
        for k in keys:
            h += f'<tr><td style="padding:6px 12px;border-bottom:1px solid {BORDER};">{k}</td>'
            for yr in YEARS:
                v = M[yr][k]
                c = _color_for(k, v)
                h += (f'<td style="padding:6px 12px;border-bottom:1px solid {BORDER};'
                      f'text-align:right;color:{c};font-weight:600;">{fmt(v)}</td>')
            h += "</tr>"
        return h

    hdr = (f'<tr><th style="padding:8px 12px;text-align:left;border-bottom:2px solid {BLUE};">Metric</th>'
           + "".join(f'<th style="padding:8px 12px;text-align:right;border-bottom:2px solid {BLUE};">{y}</th>' for y in YEARS)
           + "</tr>")
    sect_row = lambda t: (f'<tr><td colspan="{len(YEARS)+1}" style="padding:8px 12px;color:{BLUE};'
                          f'font-weight:700;background:{BG};">{t}</td></tr>')

    st.markdown(
        f'<table style="width:100%;border-collapse:collapse;background:{CARD_BG};'
        f'border-radius:10px;overflow:hidden;font-size:.88rem;">'
        f'{hdr}'
        f'{sect_row("Structural Aggregates (€M)")}{_rows(agg_keys, fm)}'
        f'{sect_row("Liquidity Ratios")}{_rows(liq_keys, fr)}'
        f'{sect_row("Solvency &amp; Leverage")}{_rows(sol_keys, fr)}'
        f'</table>',
        unsafe_allow_html=True,
    )


# =====================================================================
# PAGE 2 — Balance Sheet Explorer
# =====================================================================
def page_balance_sheet(df, M):
    st.title("📊 Balance Sheet Explorer")

    # ── Full balance sheet table ──
    section("Full Balance Sheet (€M)")
    hdr = (f'<tr><th style="padding:6px 12px;text-align:left;border-bottom:2px solid {BLUE};">Line Item</th>'
           + "".join(f'<th style="padding:6px 12px;text-align:right;border-bottom:2px solid {BLUE};">{y}</th>' for y in YEARS)
           + "</tr>")
    rows_html = ""
    for item in ALL_ITEMS:
        if item not in df.index:
            continue
        is_t = item.startswith("Total")
        bg = f"background:{BG};" if is_t else ""
        fw = "font-weight:700;" if is_t else ""
        rows_html += f'<tr style="{bg}"><td style="padding:5px 12px;border-bottom:1px solid {BORDER};{fw}">{item}</td>'
        for yr in YEARS:
            v = _g(df, item, yr) / 1000
            rows_html += (f'<td style="padding:5px 12px;border-bottom:1px solid {BORDER};'
                          f'text-align:right;{fw}">€{v:,.1f}M</td>')
        rows_html += "</tr>"

    st.markdown(
        f'<table style="width:100%;border-collapse:collapse;background:{CARD_BG};'
        f'border-radius:10px;overflow:hidden;font-size:.84rem;">{hdr}{rows_html}</table>',
        unsafe_allow_html=True,
    )

    # ── Asset composition stacked bar ──
    section("Asset Composition (€M)")
    asset_detail = [i for i in ASSET_ITEMS if not i.startswith("Total")]
    fig = go.Figure()
    for idx, item in enumerate(asset_detail):
        vals = [_g(df, item, yr) / 1000 for yr in YEARS]
        fig.add_trace(go.Bar(x=YEARS_STR, y=vals, name=sn(item),
                              marker_color=PALETTE[idx % len(PALETTE)]))
    fig.update_layout(barmode="stack", yaxis_title="€M", title="Asset Breakdown by Year")
    st.plotly_chart(dark_fig(fig), use_container_width=True)

    # ── Equity & Liabilities composition stacked bar ──
    section("Equity & Liabilities Composition (€M)")
    el_detail = [i for i in EQLIAB_ITEMS
                 if not i.startswith("Total")
                 and any(_g(df, i, yr) != 0 for yr in YEARS)]
    fig = go.Figure()
    for idx, item in enumerate(el_detail):
        vals = [_g(df, item, yr) / 1000 for yr in YEARS]
        fig.add_trace(go.Bar(x=YEARS_STR, y=vals, name=sn(item),
                              marker_color=PALETTE[idx % len(PALETTE)]))
    fig.update_layout(barmode="stack", yaxis_title="€M", title="E&L Breakdown by Year")
    st.plotly_chart(dark_fig(fig), use_container_width=True)

    # ── Structural % breakdown ──
    section("Structural Breakdown (%)")
    c1, c2 = st.columns(2)
    with c1:
        fig = go.Figure()
        nca_p = [_g(df, "Total non-current assets", yr) / _g(df, "Total assets", yr) * 100 for yr in YEARS]
        ca_p  = [_g(df, "Total current assets", yr) / _g(df, "Total assets", yr) * 100 for yr in YEARS]
        fig.add_trace(go.Bar(x=YEARS_STR, y=nca_p, name="Non-Current Assets %", marker_color=BLUE))
        fig.add_trace(go.Bar(x=YEARS_STR, y=ca_p, name="Current Assets %", marker_color=GREEN))
        fig.update_layout(barmode="stack", yaxis_title="%", title="Assets: NCA vs CA")
        st.plotly_chart(dark_fig(fig, 360), use_container_width=True)
    with c2:
        fig = go.Figure()
        for lbl, key, col in [("Equity %", "Total equity", GREEN),
                               ("NCL %", "Total non-current liabilities", BLUE),
                               ("CL %", "Total current liabilities", ORANGE)]:
            pcts = [_g(df, key, yr) / _g(df, "Total equity and liabilities", yr) * 100 for yr in YEARS]
            fig.add_trace(go.Bar(x=YEARS_STR, y=pcts, name=lbl, marker_color=col))
        fig.update_layout(barmode="stack", yaxis_title="%", title="E&L: Equity vs NCL vs CL")
        st.plotly_chart(dark_fig(fig, 360), use_container_width=True)

    # ── Detailed structural percentages table ──
    section("Each Line Item as % of Total Assets / Total E&L")
    hdr2 = (f'<tr><th style="padding:5px 12px;text-align:left;border-bottom:2px solid {BLUE};">Item</th>'
            + "".join(f'<th style="padding:5px 12px;text-align:right;border-bottom:2px solid {BLUE};">{y}</th>' for y in YEARS)
            + "</tr>")
    prows = ""
    for item in ALL_ITEMS:
        if item not in df.index or item in ("Total assets", "Total equity and liabilities"):
            continue
        denom_key = "Total assets" if item in ASSET_ITEMS else "Total equity and liabilities"
        is_t = item.startswith("Total")
        fw = "font-weight:700;" if is_t else ""
        bg = f"background:{BG};" if is_t else ""
        prows += f'<tr style="{bg}"><td style="padding:4px 12px;border-bottom:1px solid {BORDER};{fw}">{sn(item)}</td>'
        for yr in YEARS:
            d = _g(df, denom_key, yr)
            p = _g(df, item, yr) / d * 100 if d else 0
            prows += f'<td style="padding:4px 12px;border-bottom:1px solid {BORDER};text-align:right;{fw}">{p:.1f}%</td>'
        prows += "</tr>"
    st.markdown(
        f'<table style="width:100%;border-collapse:collapse;background:{CARD_BG};'
        f'border-radius:10px;overflow:hidden;font-size:.82rem;">{hdr2}{prows}</table>',
        unsafe_allow_html=True,
    )

    # ── YoY Growth table ──
    section("Year-over-Year Growth (%)")
    hdr3 = (f'<tr><th style="padding:5px 12px;text-align:left;border-bottom:2px solid {BLUE};">Item</th>'
            + "".join(f'<th style="padding:5px 12px;text-align:right;border-bottom:2px solid {BLUE};">{y}</th>' for y in YEARS[1:])
            + "</tr>")
    grows = ""
    for item in ALL_ITEMS:
        if item not in df.index:
            continue
        is_t = item.startswith("Total")
        fw = "font-weight:700;" if is_t else ""
        bg = f"background:{BG};" if is_t else ""
        grows += f'<tr style="{bg}"><td style="padding:4px 12px;border-bottom:1px solid {BORDER};{fw}">{sn(item)}</td>'
        for i in range(1, len(YEARS)):
            chg = yoy_pct(_g(df, item, YEARS[i]), _g(df, item, YEARS[i - 1]))
            grows += f'<td style="padding:4px 12px;border-bottom:1px solid {BORDER};text-align:right;{fw}">{arrow_html(chg)}</td>'
        grows += "</tr>"
    st.markdown(
        f'<table style="width:100%;border-collapse:collapse;background:{CARD_BG};'
        f'border-radius:10px;overflow:hidden;font-size:.82rem;">{hdr3}{grows}</table>',
        unsafe_allow_html=True,
    )


# =====================================================================
# PAGE 3 — Structural Aggregates
# =====================================================================
def page_aggregates(df, M):
    st.title("🏗️ Structural Aggregates")

    AGG = ["WC", "WCN", "NC", "Net Debt", "Capital Employed", "Invested Capital"]

    # ── Formula reference ──
    section("Formulas")
    st.markdown(
        f'<div class="ipanel" style="font-size:.88rem;">'
        f'<div class="formula">WC  = Total Equity + Total NCL − Total NCA</div>'
        f'<div class="formula">WCN = (Inventories + Accounts Receivable + Other CA) − (Accounts Payable + Other CL)</div>'
        f'<div class="formula">NC  = WC − WCN &nbsp;&nbsp;[verify: Cash − Borrowings Current − Bank Overdraft]</div>'
        f'<div class="formula">Net Debt = Borrowings NC + Borrowings Current − Cash</div>'
        f'<div class="formula">Capital Employed = Total NCA + WCN</div>'
        f'<div class="formula">Invested Capital = Total Equity + Net Debt</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ── Detailed cards per year ──
    section("Values by Year (€M)")
    for yr in YEARS:
        st.markdown(f"**FY {yr}**")
        cols = st.columns(6)
        for j, key in enumerate(AGG):
            v = M[yr][key]
            _, c, _ = rate_sign(v)
            with cols[j]:
                card_html(key, fm(v), c)
        # NC verification
        nc_v = M[yr]["NC_verify"]
        st.caption(f"NC verification (Cash − Borr. Current − Overdraft): {fm(nc_v)}")

    # ── Scenario classification panel ──
    section("WC / WCN / NC Scenario Classification")
    cols = st.columns(4)
    for i, yr in enumerate(YEARS):
        with cols[i]:
            scenario_card_html(yr, M[yr]["Scenario"])

    # ── Grouped bar chart of all aggregates ──
    section("Aggregates Comparison (€M)")
    fig = go.Figure()
    for idx, key in enumerate(AGG):
        vals = [M[yr][key] / 1000 for yr in YEARS]
        fig.add_trace(go.Bar(x=YEARS_STR, y=vals, name=key,
                              marker_color=PALETTE[idx % len(PALETTE)]))
    fig.update_layout(barmode="group", yaxis_title="€M")
    fig.add_hline(y=0, line_dash="dash", line_color=TXT2, opacity=0.4)
    st.plotly_chart(dark_fig(fig, 480), use_container_width=True)

    # ── YoY growth for aggregates ──
    section("Year-over-Year Growth")
    hdr = (f'<tr><th style="padding:6px 12px;text-align:left;border-bottom:2px solid {BLUE};">Aggregate</th>'
           + "".join(f'<th style="padding:6px 12px;text-align:right;border-bottom:2px solid {BLUE};">{y}</th>' for y in YEARS[1:])
           + "</tr>")
    rows_h = ""
    for key in AGG:
        rows_h += f'<tr><td style="padding:5px 12px;border-bottom:1px solid {BORDER};font-weight:600;">{key}</td>'
        for i in range(1, len(YEARS)):
            chg = yoy_pct(M[YEARS[i]][key], M[YEARS[i - 1]][key])
            rows_h += f'<td style="padding:5px 12px;border-bottom:1px solid {BORDER};text-align:right;">{arrow_html(chg)}</td>'
        rows_h += "</tr>"
    st.markdown(
        f'<table style="width:100%;border-collapse:collapse;background:{CARD_BG};'
        f'border-radius:10px;overflow:hidden;font-size:.86rem;">{hdr}{rows_h}</table>',
        unsafe_allow_html=True,
    )


# =====================================================================
# PAGE 4 — Liquidity Analysis
# =====================================================================
def _liquidity_chart(M, key, title, thresholds):
    """Bar chart for a single liquidity ratio with horizontal threshold lines."""
    fig = go.Figure()
    colors = []
    for yr in YEARS:
        v = M[yr][key]
        if key == "Current Ratio":   _, c, _ = rate_cr(v)
        elif key == "Quick Ratio":   _, c, _ = rate_qr(v)
        else:                        _, c, _ = rate_cashr(v)
        colors.append(c)
    vals = [M[yr][key] for yr in YEARS]
    fig.add_trace(go.Bar(x=YEARS_STR, y=vals, marker_color=colors, name=title,
                          text=[fr(v) for v in vals], textposition="outside",
                          textfont=dict(color=TXT)))
    for level, lbl, col, dash in thresholds:
        fig.add_hline(y=level, line_dash=dash, line_color=col, opacity=0.6,
                       annotation_text=lbl, annotation_position="top left",
                       annotation_font_color=col, annotation_font_size=11)
    fig.update_layout(yaxis_title="Ratio", title=title, showlegend=False)
    return dark_fig(fig)


def page_liquidity(df, M):
    st.title("💧 Liquidity Analysis")

    # Current Ratio
    section("Current Ratio")
    c1, c2 = st.columns([1, 2])
    with c1:
        v = M[2024]["Current Ratio"]
        e, co, sub = rate_cr(v)
        card_html("Current Ratio (2024)", f"{e} {fr(v)}", co, sub)
        interp_panel_html(
            "Current Ratio — Interpretation",
            "Current Ratio = Total Current Assets / Total Current Liabilities",
            fr(v),
            (f'<span style="color:{RED};">● &lt; 1.0 — Liquidity concern</span><br>'
             f'<span style="color:{GREEN};">● 1.5 – 2.0 — Healthy</span><br>'
             f'<span style="color:{ORANGE};">● &gt; 3.0 — Inefficient</span>'),
            f"Adidas 2024: {fr(v)} — {sub}", co,
        )
    with c2:
        thresholds = [(1, "1.0 Critical", RED, "dash"), (1.5, "1.5 Healthy min", GREEN, "dot"),
                      (2, "2.0 Healthy max", GREEN, "dot")]
        st.plotly_chart(_liquidity_chart(M, "Current Ratio", "Current Ratio", thresholds),
                        use_container_width=True)

    # Quick Ratio
    section("Quick Ratio")
    c1, c2 = st.columns([1, 2])
    with c1:
        v = M[2024]["Quick Ratio"]
        e, co, sub = rate_qr(v)
        card_html("Quick Ratio (2024)", f"{e} {fr(v)}", co, sub)
        interp_panel_html(
            "Quick Ratio — Interpretation",
            "Quick Ratio = (Total Current Assets − Inventories) / Total Current Liabilities",
            fr(v),
            (f'<span style="color:{RED};">● &lt; 1.0 — Liquidity risk</span><br>'
             f'<span style="color:{GREEN};">● 1.0 – 1.5 — Strong</span><br>'
             f'<span style="color:{ORANGE};">● &gt; 2.0 — Very conservative</span>'),
            f"Adidas 2024: {fr(v)} — {sub}", co,
        )
    with c2:
        thresholds = [(1, "1.0 Risk", RED, "dash"), (1.5, "1.5 Strong", GREEN, "dot")]
        st.plotly_chart(_liquidity_chart(M, "Quick Ratio", "Quick Ratio", thresholds),
                        use_container_width=True)

    # Cash Ratio
    section("Cash Ratio")
    c1, c2 = st.columns([1, 2])
    with c1:
        v = M[2024]["Cash Ratio"]
        e, co, sub = rate_cashr(v)
        card_html("Cash Ratio (2024)", f"{e} {fr(v)}", co, sub)
        interp_panel_html(
            "Cash Ratio — Interpretation",
            "Cash Ratio = Cash and Cash Equivalents / Total Current Liabilities",
            fr(v),
            (f'<span style="color:{ORANGE};">● &lt; 0.5 — Relies on inventory/receivables</span><br>'
             f'<span style="color:{GREEN};">● 0.5 – 1.0 — Ideal</span><br>'
             f'<span style="color:{ORANGE};">● &gt; 1.0 — Extremely conservative</span>'),
            f"Adidas 2024: {fr(v)} — {sub}", co,
        )
    with c2:
        thresholds = [(0.5, "0.5 Min ideal", ORANGE, "dash"), (1, "1.0 Max ideal", GREEN, "dot")]
        st.plotly_chart(_liquidity_chart(M, "Cash Ratio", "Cash Ratio", thresholds),
                        use_container_width=True)

    # Comparative chart
    section("Comparative — All Liquidity Ratios")
    fig = go.Figure()
    for name, key, col in [("Current Ratio", "Current Ratio", BLUE),
                            ("Quick Ratio", "Quick Ratio", GREEN),
                            ("Cash Ratio", "Cash Ratio", ORANGE)]:
        vals = [M[yr][key] for yr in YEARS]
        fig.add_trace(go.Scatter(x=YEARS_STR, y=vals, name=name, mode="lines+markers",
                                  line=dict(color=col, width=3), marker=dict(size=9)))
    fig.add_hline(y=1, line_dash="dash", line_color=RED, opacity=0.4,
                   annotation_text="1.0", annotation_font_color=RED)
    fig.update_layout(yaxis_title="Ratio", xaxis_title="Year")
    st.plotly_chart(dark_fig(fig), use_container_width=True)


# =====================================================================
# PAGE 5 — Solvency & Leverage
# =====================================================================
def _solvency_chart(M, key, title, thresholds):
    fig = go.Figure()
    colors = []
    for yr in YEARS:
        v = M[yr][key]
        if key == "Total Debt Ratio": _, c, _ = rate_tdr(v)
        elif key == "Debt-to-Equity": _, c, _ = rate_de(v)
        else:                         c = BLUE
        colors.append(c)
    vals = [M[yr][key] for yr in YEARS]
    fig.add_trace(go.Bar(x=YEARS_STR, y=vals, marker_color=colors, name=title,
                          text=[fr(v) for v in vals], textposition="outside",
                          textfont=dict(color=TXT)))
    for level, lbl, col, dash in thresholds:
        fig.add_hline(y=level, line_dash=dash, line_color=col, opacity=0.6,
                       annotation_text=lbl, annotation_position="top left",
                       annotation_font_color=col, annotation_font_size=11)
    fig.update_layout(yaxis_title="Ratio", title=title, showlegend=False)
    return dark_fig(fig)


def page_solvency(df, M):
    st.title("🏦 Solvency & Leverage")

    # Equity Multiplier
    section("Equity Multiplier")
    c1, c2 = st.columns([1, 2])
    with c1:
        v = M[2024]["Equity Multiplier"]
        card_html("Equity Multiplier (2024)", fr(v), BLUE,
                  f"Higher = more leveraged. Adidas uses €{fr(v)} of assets per €1 of equity.")
        interp_panel_html(
            "Equity Multiplier — Interpretation",
            "Equity Multiplier = Total Assets / Total Equity",
            fr(v),
            (f'<span style="color:{TXT};">Measures financial leverage aggressiveness.<br>'
             f'EM = 1 → all-equity financed. Higher → more debt-financed.</span>'),
            f"Adidas 2024: EM of {fr(v)} means for every €1 of equity, the company controls "
            f"€{fr(v)} of assets — moderate-to-high leverage.", BLUE,
        )
    with c2:
        st.plotly_chart(_solvency_chart(M, "Equity Multiplier", "Equity Multiplier", []),
                        use_container_width=True)

    # Total Debt Ratio
    section("Total Debt Ratio")
    c1, c2 = st.columns([1, 2])
    with c1:
        v = M[2024]["Total Debt Ratio"]
        e, co, sub = rate_tdr(v)
        card_html("Total Debt Ratio (2024)", f"{e} {fr(v)}", co, sub)
        interp_panel_html(
            "Total Debt Ratio — Interpretation",
            "Total Debt Ratio = (Borrowings NC + Borrowings Current) / Total Assets",
            fr(v),
            (f'<span style="color:{GREEN};">● 0.2 – 0.4 — Low debt</span><br>'
             f'<span style="color:{ORANGE};">● 0.4 – 0.6 — Acceptable</span><br>'
             f'<span style="color:{RED};">● &gt; 0.6 — High</span><br>'
             f'<span style="color:{RED};">● &gt; 0.8 — Very high</span>'),
            f"Adidas 2024: {fr(v)} — {sub}", co,
        )
    with c2:
        thresholds = [(0.4, "0.4 Low max", GREEN, "dot"), (0.6, "0.6 Acceptable max", ORANGE, "dash"),
                      (0.8, "0.8 High", RED, "dash")]
        st.plotly_chart(_solvency_chart(M, "Total Debt Ratio", "Total Debt Ratio", thresholds),
                        use_container_width=True)

    # Debt-to-Equity
    section("Debt-to-Equity")
    c1, c2 = st.columns([1, 2])
    with c1:
        v = M[2024]["Debt-to-Equity"]
        e, co, sub = rate_de(v)
        card_html("Debt-to-Equity (2024)", f"{e} {fr(v)}", co, sub)
        interp_panel_html(
            "Debt-to-Equity — Interpretation",
            "D/E = (Borrowings NC + Borrowings Current) / Total Equity",
            fr(v),
            (f'<span style="color:{GREEN};">● 0.5 – 1.0 — Healthy</span><br>'
             f'<span style="color:{ORANGE};">● 1.0 – 1.5 — Acceptable</span><br>'
             f'<span style="color:{RED};">● &gt; 1.5 — High leverage</span>'),
            f"Adidas 2024: {fr(v)} — {sub}", co,
        )
    with c2:
        thresholds = [(1, "1.0 Healthy max", GREEN, "dot"), (1.5, "1.5 Acceptable max", ORANGE, "dash")]
        st.plotly_chart(_solvency_chart(M, "Debt-to-Equity", "Debt-to-Equity", thresholds),
                        use_container_width=True)

    # Comparative chart
    section("Comparative — Solvency Ratios")
    fig = go.Figure()
    for name, key, col in [("Equity Multiplier", "Equity Multiplier", BLUE),
                            ("Total Debt Ratio", "Total Debt Ratio", ORANGE),
                            ("Debt-to-Equity", "Debt-to-Equity", GREEN)]:
        vals = [M[yr][key] for yr in YEARS]
        fig.add_trace(go.Scatter(x=YEARS_STR, y=vals, name=name, mode="lines+markers",
                                  line=dict(color=col, width=3), marker=dict(size=9)))
    fig.update_layout(yaxis_title="Ratio", xaxis_title="Year")
    st.plotly_chart(dark_fig(fig), use_container_width=True)


# =====================================================================
# PAGE 6 — Interactive Comparison Tool
# =====================================================================
def page_comparison(df, M):
    st.title("🔀 Interactive Comparison Tool")

    # Build master metrics catalogue
    eur_keys = ["WC", "WCN", "NC", "Net Debt", "Capital Employed", "Invested Capital"]
    ratio_keys = ["Current Ratio", "Quick Ratio", "Cash Ratio",
                  "Equity Multiplier", "Total Debt Ratio", "Debt-to-Equity"]
    raw_items_available = [i for i in ALL_ITEMS if i in df.index]

    # ── Controls ──
    c1, c2 = st.columns(2)
    with c1:
        sel_years = st.multiselect("Select Years", YEARS, default=YEARS)
    with c2:
        metric_group = st.selectbox("Metric Group",
                                     ["Structural Aggregates (€M)", "Liquidity & Solvency Ratios",
                                      "Raw Balance Sheet Items (€M)"])
    if not sel_years:
        st.warning("Select at least one year.")
        return

    if metric_group == "Structural Aggregates (€M)":
        avail = eur_keys
        fmt = fm
        unit = "€M"
    elif metric_group == "Liquidity & Solvency Ratios":
        avail = ratio_keys
        fmt = fr
        unit = "Ratio"
    else:
        avail = raw_items_available
        fmt = fm
        unit = "€M"

    sel_metrics = st.multiselect("Select Metrics", avail, default=avail[:4])
    if not sel_metrics:
        st.warning("Select at least one metric.")
        return

    # ── Dynamic chart ──
    section("Comparison Chart")
    fig = go.Figure()
    for idx, key in enumerate(sel_metrics):
        if key in eur_keys or key in ratio_keys:
            vals = [M[yr][key] / 1000 if key in eur_keys else M[yr][key] for yr in sel_years]
        else:
            vals = [_g(df, key, yr) / 1000 for yr in sel_years]
        fig.add_trace(go.Bar(
            x=[str(y) for y in sel_years], y=vals, name=sn(key),
            marker_color=PALETTE[idx % len(PALETTE)],
        ))
    fig.update_layout(barmode="group", yaxis_title=unit, xaxis_title="Year")
    st.plotly_chart(dark_fig(fig, 450), use_container_width=True)

    # ── Dynamic table with YoY deltas ──
    section("Values & Year-over-Year Change")
    yr_strs = [str(y) for y in sel_years]
    hdr = f'<tr><th style="padding:6px 12px;text-align:left;border-bottom:2px solid {BLUE};">Metric</th>'
    for y in sel_years:
        hdr += f'<th style="padding:6px 12px;text-align:right;border-bottom:2px solid {BLUE};">{y}</th>'
    # Add delta columns
    for i in range(1, len(sel_years)):
        hdr += f'<th style="padding:6px 12px;text-align:right;border-bottom:2px solid {BLUE};">Δ {sel_years[i-1]}→{sel_years[i]}</th>'
    hdr += "</tr>"

    trows = ""
    for key in sel_metrics:
        trows += f'<tr><td style="padding:5px 12px;border-bottom:1px solid {BORDER};font-weight:600;">{sn(key)}</td>'
        vals = []
        for yr in sel_years:
            if key in eur_keys or key in ratio_keys:
                v = M[yr][key]
            else:
                v = _g(df, key, yr)
            vals.append(v)
            disp = fmt(v)
            trows += f'<td style="padding:5px 12px;border-bottom:1px solid {BORDER};text-align:right;">{disp}</td>'
        for i in range(1, len(sel_years)):
            chg = yoy_pct(vals[i], vals[i - 1])
            trows += f'<td style="padding:5px 12px;border-bottom:1px solid {BORDER};text-align:right;">{arrow_html(chg)}</td>'
        trows += "</tr>"

    st.markdown(
        f'<table style="width:100%;border-collapse:collapse;background:{CARD_BG};'
        f'border-radius:10px;overflow:hidden;font-size:.86rem;">{hdr}{trows}</table>',
        unsafe_allow_html=True,
    )


# =====================================================================
# PAGE 7 — Trend & Forecast
# =====================================================================
def page_trends(df, M):
    st.title("📈 Trend & Forecast")
    st.caption("Linear regression on 4 data points (2021-2024). Projections are indicative only.")

    forecast_keys = [
        ("WC", "Working Capital", fm, True),
        ("NC", "Net Cash", fm, True),
        ("Current Ratio", "Current Ratio", fr, False),
        ("Net Debt", "Net Debt", fm, True),
        ("Debt-to-Equity", "Debt-to-Equity", fr, False),
    ]

    x = np.array([0, 1, 2, 3], dtype=float)
    x_proj = np.array([0, 1, 2, 3, 4], dtype=float)
    proj_year = 2025
    years_ext = YEARS + [proj_year]
    years_ext_str = [str(y) for y in years_ext]

    for key, title, fmt, is_eur in forecast_keys:
        section(f"{title} — Trend & 2025 Projection")
        y = np.array([M[yr][key] for yr in YEARS], dtype=float)
        if is_eur:
            y_display = y / 1000  # €M
        else:
            y_display = y.copy()

        # Linear regression
        coeffs = np.polyfit(x, y_display, 1)
        y_fit = np.polyval(coeffs, x)
        y_proj_val = np.polyval(coeffs, 4)
        y_fit_ext = np.polyval(coeffs, x_proj)

        # R²
        ss_res = np.sum((y_display - y_fit) ** 2)
        ss_tot = np.sum((y_display - np.mean(y_display)) ** 2)
        r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0

        c1, c2 = st.columns([1, 2])
        with c1:
            unit = "€M" if is_eur else ""
            card_html(f"2025 Projection", f"{fmt(y_proj_val * 1000) if is_eur else fr(y_proj_val)}",
                      BLUE, f"R² = {r2:.3f}")
            if r2 < 0.5:
                st.markdown(
                    f'<div class="card" style="border-left:3px solid {ORANGE};">'
                    f'<span style="color:{ORANGE};font-weight:600;">⚠️ Low R² ({r2:.3f})</span><br>'
                    f'<span style="font-size:.82rem;color:{TXT2};">Trend explains &lt;50% of variance. '
                    f'Projection is unreliable — data is heavily influenced by other factors.</span></div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f'<div class="card" style="border-left:3px solid {GREEN};">'
                    f'<span style="color:{GREEN};font-weight:600;">✅ R² = {r2:.3f}</span><br>'
                    f'<span style="font-size:.82rem;color:{TXT2};">Trend line fits the data reasonably well.</span></div>',
                    unsafe_allow_html=True,
                )

        with c2:
            fig = go.Figure()
            # Actual data points
            fig.add_trace(go.Scatter(
                x=YEARS_STR, y=y_display.tolist(), name="Actual", mode="lines+markers",
                line=dict(color=BLUE, width=3), marker=dict(size=10),
            ))
            # Trend line + projection
            fig.add_trace(go.Scatter(
                x=years_ext_str, y=y_fit_ext.tolist(), name="Trend", mode="lines",
                line=dict(color=GREEN, width=2, dash="dash"),
            ))
            # Projection point
            fig.add_trace(go.Scatter(
                x=[str(proj_year)], y=[y_proj_val], name=f"2025 Projection",
                mode="markers", marker=dict(size=14, color=ORANGE, symbol="diamond"),
            ))
            unit_label = "€M" if is_eur else "Ratio"
            fig.update_layout(yaxis_title=unit_label, xaxis_title="Year", title=title)
            st.plotly_chart(dark_fig(fig), use_container_width=True)


# =====================================================================
# MAIN
# =====================================================================
PAGES = {
    "📋 Executive Summary": page_summary,
    "📊 Balance Sheet Explorer": page_balance_sheet,
    "🏗️ Structural Aggregates": page_aggregates,
    "💧 Liquidity Analysis": page_liquidity,
    "🏦 Solvency & Leverage": page_solvency,
    "🔀 Interactive Comparison": page_comparison,
    "📈 Trend & Forecast": page_trends,
}


def main():
    inject_css()

    # Sidebar
    with st.sidebar:
        st.markdown(
            f'<div style="text-align:center;padding:12px 0;">'
            f'<div style="font-size:2rem;">👟</div>'
            f'<div style="font-size:1.2rem;font-weight:700;color:{TXT};">Adidas</div>'
            f'<div style="font-size:.75rem;color:{TXT2};">Financial Analysis Dashboard</div>'
            f'<div style="font-size:.65rem;color:{TXT2};margin-top:4px;">Balance Sheet 2021 – 2024</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        st.markdown("---")
        page = st.radio("Navigation", list(PAGES.keys()), label_visibility="collapsed")

    # Load data + compute
    if not FILE_PATH.exists():
        st.error(f"Data file not found: {FILE_PATH}")
        return

    df = load_data()
    M = compute(df)

    # Route to selected page
    PAGES[page](df, M)


if __name__ == "__main__":
    main()
